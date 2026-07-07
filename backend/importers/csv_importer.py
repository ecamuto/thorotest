import csv
import io
from .base import ImportResult, TestData

# Priority normalisation
_PRIORITY_MAP = {
    "critical": "critical", "blocker": "critical",
    "high": "high", "major": "high",
    "medium": "med", "med": "med", "normal": "med",
    "low": "low", "minor": "low", "trivial": "low",
    # TestRail numeric priority IDs
    "1": "low", "2": "med", "3": "high", "4": "critical",
}

# Type normalisation
_TYPE_MAP = {
    "automated": "automated", "automatic": "automated", "auto": "automated",
    "manual": "manual",
    "1": "automated",   # TestRail type_id 1 = Automated
}


def _normalise_priority(v: str) -> str:
    return _PRIORITY_MAP.get(v.strip().lower(), "med")


def _normalise_type(v: str) -> str:
    return _TYPE_MAP.get(v.strip().lower(), "manual")


# Column alias tables: canonical_name → list of source column headers (case-insensitive)
_COLUMN_ALIASES = {
    "title":       ["title", "name", "test name", "case name", "summary"],
    "folder_path": ["section", "folder", "suite", "component", "area path", "folder path", "section hierarchy"],
    "type":        ["type", "type_id", "test type", "automated test name"],
    "status":      ["status", "state", "result"],
    "priority":    ["priority", "priority_id", "severity"],
    "owner":       ["owner", "assigned to", "assignee"],
    "tags":        ["tags", "labels", "keywords", "categories"],
    "source_id":   ["id", "test id", "case id", "work item id"],
}


def _detect_columns(headers: list[str]) -> dict[str, str]:
    """Return mapping canonical_name → actual_header for detected columns."""
    lower_headers = {h.strip().lower(): h for h in headers}
    mapping = {}
    for canonical, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in lower_headers:
                mapping[canonical] = lower_headers[alias]
                break
    return mapping


def _detect_tool(headers: list[str]) -> str:
    lower = {h.strip().lower() for h in headers}
    if "section" in lower or "section hierarchy" in lower:
        return "testrail"
    if "component" in lower or "sprint" in lower:
        return "zephyr"
    if "work item type" in lower or "area path" in lower:
        return "azure"
    return "generic"


def _result_from_rows(headers: list[str], rows, column_mapping: dict | None,
                      label: str) -> ImportResult:
    """Build an ImportResult from tabular rows (list of {header: value} dicts).

    Shared by the CSV and XLSX importers. `label` is the format prefix
    ("csv" / "xlsx") used in format_detected.
    """
    auto_mapping = _detect_columns(headers)
    mapping = {**auto_mapping, **(column_mapping or {})}

    tool = _detect_tool(headers)
    tests: list[TestData] = []
    warnings: list[str] = []

    if "title" not in mapping:
        warnings.append("No title column detected — cannot import test cases from this file")
        return ImportResult(format_detected=f"{label} ({tool})", warnings=warnings,
                            source_provider=tool)

    for row in rows:
        title = str(row.get(mapping["title"], "") or "").strip()
        if not title:
            continue

        folder_path = ""
        if "folder_path" in mapping:
            folder_path = str(row.get(mapping["folder_path"], "") or "").strip()
            # Azure uses " > " separator, TestRail uses " / " — normalise to "/"
            folder_path = folder_path.replace(" > ", "/").strip("/")

        raw_priority = str(row.get(mapping.get("priority", ""), "") or "").strip()
        raw_type = str(row.get(mapping.get("type", ""), "") or "").strip()

        raw_tags = str(row.get(mapping.get("tags", ""), "") or "").strip()
        tags = [t.strip() for t in raw_tags.replace(";", ",").split(",") if t.strip()] if raw_tags else []

        tests.append(TestData(
            title=title,
            folder_path=folder_path,
            type=_normalise_type(raw_type) if raw_type else "manual",
            status="pending",
            priority=_normalise_priority(raw_priority) if raw_priority else "med",
            owner=str(row.get(mapping.get("owner", ""), "") or "").strip(),
            tags=tags,
            source_id=str(row.get(mapping.get("source_id", ""), "") or "").strip(),
        ))

    return ImportResult(
        tests=tests,
        format_detected=f"{label} ({tool})",
        warnings=warnings,
        source_provider=tool,
    )


def _csv_rows(content: bytes) -> tuple[list[str], list[dict]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader.fieldnames or []), list(reader)


def _xlsx_rows(content: bytes) -> tuple[list[str], list[dict]]:
    """Read the first worksheet of an .xlsx into (headers, row dicts)."""
    from openpyxl import load_workbook  # lazy: only needed for .xlsx imports

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        rows.append({headers[i]: values[i] for i in range(min(len(headers), len(values)))})
    wb.close()
    return headers, rows


def parse_csv(content: bytes, column_mapping: dict | None = None) -> ImportResult:
    """
    Parse a CSV export from TestRail, Zephyr Scale, Azure Test Plans, or generic CSV.
    column_mapping: optional override {canonical_name: actual_header}.
    """
    headers, rows = _csv_rows(content)
    return _result_from_rows(headers, rows, column_mapping, "csv")


def parse_xlsx(content: bytes, column_mapping: dict | None = None) -> ImportResult:
    """Parse an .xlsx export (first worksheet), reusing the CSV column logic."""
    try:
        headers, rows = _xlsx_rows(content)
    except Exception as e:
        return ImportResult(format_detected="xlsx",
                            warnings=[f"Could not read .xlsx file: {e}"])
    return _result_from_rows(headers, rows, column_mapping, "xlsx")


def get_csv_columns(content: bytes) -> dict:
    """Return detected headers and column mapping for UI display."""
    headers, _ = _csv_rows(content)
    return {"headers": headers, "mapping": _detect_columns(headers), "tool": _detect_tool(headers)}


def get_xlsx_columns(content: bytes) -> dict:
    """Return detected headers and column mapping for UI display (.xlsx)."""
    try:
        headers, _ = _xlsx_rows(content)
    except Exception:
        headers = []
    return {"headers": headers, "mapping": _detect_columns(headers), "tool": _detect_tool(headers)}
